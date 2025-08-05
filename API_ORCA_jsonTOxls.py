import json
import pandas as pd
import requests
import openpyxl
from openpyxl import Workbook
from pprint import pprint
import urllib.parse
import os

# ______________________________________________________________________________________________________

raw_data = {"query":{"models":["Alert"],"type":"object_set","with":{"operator":"and","type":"operation","values":[{"key":"Status","values":["open","in_progress"],"type":"str","operator":"in"},{"key":"RiskLevel","values":["critical","high"],"type":"str","operator":"in"}]}},"standard_format":False,"with_model_names":True,"get_results_and_count":True,"order_by_pk":False,"additional_models[]":["CloudAccount","CodeOrigins","Inventory"],"enable_pagination":True,"limit":1000,"start_at_index":0,"order_by[]":["-OrcaScore"],"max_tier":2}

headers = {"Authorization": "Token "}
#response = requests.get("https://api.orcasecurity.io/api/alerts?state.risk_level=critical,high", headers=headers)
response = requests.post("https://api.orcasecurity.io/api/sonar/query", headers=headers, json=raw_data)

if response.status_code == 200:
    new_data = response.json()

    try:
        with open("data1.json", "r") as json_file:
            existing_data = json.load(json_file)
    except (FileNotFoundError, json.decoder.JSONDecodeError):
        existing_data = []

    existing_data.append(new_data)

    with open("data1.json", "w") as json_file:
        json.dump(existing_data, json_file, indent=4)
        print("Arquivo data.json criado com sucesso!")
else:
    print("Failed to retrieve data from the API. Status code:", response.status_code)
# ________________________________________________________________________________________________________

json_file = "data1.json"
xlsx_file = "data1.xlsx"

with open(json_file, "r", encoding="utf-8") as file:
    json_data = json.load(file)

records = json_data[0].get("data", [])  # Pegamos a lista de registros
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Alertas"
headers = ["Nome_Conta", "Titulo", "Descricao", "Severidade"]
ws.append(headers)

for record in records:
    ws.append([
        record.get("data", {}).get("CloudAccount", {}).get("name", "N/A"),
        record.get("data", {}).get("AlertType", {}).get("value", "N/A"),
        record.get("data", {}).get("Inventory", {}).get("name", "N/A"),
        record.get("data", {}).get("RiskLevel", {}).get("value", "N/A")
    ])

wb.save(xlsx_file)
print(f"Arquivo {xlsx_file} criado com sucesso!")

#____________________________________________________________________________________________________________

raw_data = {"query":{"models":["Alert"],"type":"object_set","with":{"operator":"and","type":"operation","values":[{"key":"Status","values":["open","in_progress"],"type":"str","operator":"in"},{"key":"RiskLevel","values":["critical","high"],"type":"str","operator":"in"}]}},"standard_format":False,"with_model_names":True,"get_results_and_count":True,"order_by_pk":False,"additional_models[]":["CloudAccount","CodeOrigins","Inventory"],"enable_pagination":True,"limit":1000,"start_at_index":1000,"order_by[]":["-OrcaScore"],"max_tier":2}

headers = {"Authorization": "Token "}
#response = requests.get("https://api.orcasecurity.io/api/alerts?state.risk_level=critical,high", headers=headers)
response = requests.post("https://api.orcasecurity.io/api/sonar/query", headers=headers, json=raw_data)

if response.status_code == 200:
    new_data = response.json()

    try:
        with open("data2.json", "r") as json_file:
            existing_data = json.load(json_file)
    except (FileNotFoundError, json.decoder.JSONDecodeError):
        existing_data = []

    existing_data.append(new_data)

    with open("data2.json", "w") as json_file:
        json.dump(existing_data, json_file, indent=4)
        print("Arquivo data.json criado com sucesso!")
else:
    print("Failed to retrieve data from the API. Status code:", response.status_code)
# ________________________________________________________________________________________________________

json_file = "data2.json"
xlsx_file = "data2.xlsx"

with open(json_file, "r", encoding="utf-8") as file:
    json_data = json.load(file)

records = json_data[0].get("data", [])  # Pegamos a lista de registros
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Alertas"
headers = ["Nome_Conta", "Titulo", "Descricao", "Severidade"]
ws.append(headers)

for record in records:
    ws.append([
        record.get("data", {}).get("CloudAccount", {}).get("name", "N/A"),
        record.get("data", {}).get("AlertType", {}).get("value", "N/A"),
        record.get("data", {}).get("Inventory", {}).get("name", "N/A"),
        record.get("data", {}).get("RiskLevel", {}).get("value", "N/A")
    ])

wb.save(xlsx_file)
print(f"Arquivo {xlsx_file} criado com sucesso!")

df1 = pd.read_excel("data1.xlsx")
df2 = pd.read_excel("data2.xlsx")
df_concatenado = pd.concat([df1, df2], ignore_index=True)
df_concatenado.to_excel("data3.xlsx", index=False)
print(f"Arquivo data3.xlsx criado com sucesso!")

# ______________________________________________________________________________________________________

raw_data = {
        "current_page": 1,
        "page_size": 1000,
        "sort": {
            "sort_by": "creation_time",
            "sort_order": "desc"
        },
        "show_feature_info": True
    }

headers = {"Authorization": "Token "}
response = requests.post("https://api.orcasecurity.io/api/accountcenter/accounts", headers=headers, json=raw_data)

if response.status_code == 200:
    new_data = response.json()

    try:
        with open("data4.json", "r") as json_file:
            existing_data = json.load(json_file)
    except (FileNotFoundError, json.decoder.JSONDecodeError):
        existing_data = []

    existing_data.append(new_data)

    with open("data4.json", "w") as json_file:
        json.dump(existing_data, json_file, indent=4)
        print("Arquivo data4.json criado com sucesso!")
else:
    print("Failed to retrieve data from the API. Status code:", response.status_code)
# ________________________________________________________________________________________________________

json_file = "data4.json"
xlsx_file = "data4.xlsx"

with open(json_file, "r", encoding="utf-8") as file:
    json_data = json.load(file)

records = json_data[0].get("accounts", [])  # Pegamos a lista de registros
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Accounts"
headers = ["Nome_Conta", "Cloud_Provider", "Produtos", "Workloads", "Status"]
ws.append(headers)

for record in records:
    business_units = record.get("business_units") or []
    if isinstance(business_units, list):
        business_units_names = ", ".join(unit.get("name", "") for unit in business_units if isinstance(unit, dict))
    else:
        business_units_names = ""
    workloads_raw = record.get("workloads_count", {}).get("workloads_number", 0)
    try:
        workloads_number = int(workloads_raw)
    except (ValueError, TypeError):
        workloads_number = 0  # ou qualquer valor padrão desejado
    ws.append([
        record.get("account_name", "N/A"),
        record.get("cloud_provider", "N/A"),
        business_units_names,
        workloads_number,
        record.get("status", "N/A")
    ])

wb.save(xlsx_file)
print(f"Arquivo {xlsx_file} criado com sucesso!")

xlsx_file = "data1.xlsx"

if os.path.exists(xlsx_file):
    os.remove(xlsx_file)
    print(f"Arquivo {xlsx_file} foi excluído com sucesso.")
else:
    print(f"O arquivo {xlsx_file} não existe.")

xlsx_file = "data2.xlsx"

if os.path.exists(xlsx_file):
    os.remove(xlsx_file)
    print(f"Arquivo {xlsx_file} foi excluído com sucesso.")
else:                                                                                                                       
    print(f"O arquivo {xlsx_file} não existe.")
